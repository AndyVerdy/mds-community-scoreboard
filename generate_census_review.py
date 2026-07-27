import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Colors
HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NEW_FILL = PatternFill(start_color='00E676', end_color='00E676', fill_type='solid')
NEW_FONT = Font(bold=True)
CAT_FILL = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
ALT_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
BORDER = Border(
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC')
)
PURPOSE_FONT = Font(color='1565C0')
WHY_FONT = Font(color='2E7D32')
FORM_FONT = Font(bold=True, color='E65100')
WRAP = Alignment(wrap_text=True, vertical='top')

# Master data: [category, topic, f1Ref, f1Text, f1Subtext, f2Ref, f2Text, f2Subtext, f3Ref, f3Text, f3Subtext, purpose, why, whichForm, isNew]
master_rows = [
    # IDENTITY
    ['Identity', 'Full name', 'F1-Q1', 'What is your full name?', '', 'F2-Q1', 'What is your full name?', '', 'F3-Q1', 'What is your full name?', '',
     'Member identification', 'Core identity field needed to match records across systems, CRM, and communications', 'All 3 forms', False],
    ['Identity', 'Email', 'F1-Q2', 'What is your email address?', '', 'F2-Q2', 'Email address', '', 'F3-Q2', 'What is your email address?', '',
     'Primary contact & record matching', 'Unique identifier to link census responses to member records in Airtable/CRM. Essential for data integrity', 'All 3 forms', False],
    ['Identity', 'Formal title', 'F1-Q18', 'What is your formal title in your organization?', '', '', '', '', 'F3-Q34', 'What is your formal title within your organization?', '',
     'Understand member seniority/role', 'Helps segment members by role (CEO vs COO vs Manager). Useful for matching members in peer groups and understanding decision-making authority', 'Application + Standard Census', False],
    ['Identity', 'Day-to-day responsibilities', 'F1-Q19', 'What are your day-to-day responsibilities?', '', '', '', '', 'F3-Q35', 'What are your day to day responsibilities?', '',
     'Understand what members actually do', 'Title alone doesnt tell the full story. Reveals if a CEO is hands-on in ops vs strategy, helping match members for relevant conversations', 'Application + Standard Census', False],
    ['Identity', 'Education level', '', '', '', '', '', '', 'F3-Q3', 'Highest level of education', '',
     'Demographic segmentation', 'Background context for member profiles. Can correlate with business outcomes. Low priority - only ask once at application', 'Application only', False],
    ['Identity', 'Prior work before e-commerce', '', '', '', '', '', '', 'F3-Q5', 'What did you do before you started selling on Amazon?', '',
     'Understand member background', 'One-time question. Helps find members with shared professional backgrounds. Doesnt change - no need to re-ask', 'Application only', False],
    ['Identity', 'Gsuite email for doc access', '', '', '', 'F2-Q32', 'What is your Gsuite email?', 'For MDS shared doc access', 'F3-Q48', 'What is your gsuite email?', '',
     'Grant access to shared Google resources', 'Needed to add members to shared Google Drives, Docs, and Sheets. Different from primary email if they use Gsuite for business', 'Application + Standard Census', False],

    # NEW: BRAND INFO
    ['Brand Info', 'Business name', '', '', '', '', '', '', '', '', '',
     'Official business identification', 'Need to know the legal/DBA business name for verification, directory listings, and matching with revenue data. Different from brand name', 'Application + Standard Census', True],
    ['Brand Info', 'Website URL', '', '', '', '', '', '', '', '', '',
     'Verify business legitimacy & size', 'Cross-reference with application claims. Enables staff to quickly verify a members business. Useful for member directory', 'Application + Standard Census', True],
    ['Brand Info', 'Amazon storefront name/URL', '', '', '', '', '', '', '', '', '',
     'Verify Amazon presence & revenue', 'Primary verification tool. Staff can check storefront to validate revenue claims, product count, and brand registry status', 'Application + Standard Census', True],
    ['Brand Info', 'TikTok shop name/code', '', '', '', '', '', '', '', '', '',
     'Track TikTok commerce adoption', 'Growing channel for e-com sellers. Helps understand member diversification and enables TikTok-specific peer matching', 'Application + Standard Census', True],
    ['Brand Info', 'Revenue screenshot upload', '', '', '', '', '', '', '', '', '',
     'Verify reported revenue figures', 'Hard evidence to validate self-reported revenue. Reduces false claims and maintains group credibility. Ask annually for re-verification', 'Application + Standard Census', True],
    ['Brand Info', 'TikTok revenue (annual)', '', '', '', '', '', '', '', '', '',
     'Track channel-specific revenue', 'TikTok Shop is fastest growing channel. Separate field gives cleaner data than trying to extract from a matrix. Complements the channel matrix', 'Application + Standard Census', True],
    ['Brand Info', 'Shopify revenue (annual)', '', '', '', '', '', '', '', '', '',
     'Track DTC revenue separately', 'Shopify/DTC revenue is a key indicator of brand strength vs Amazon dependency. Important for segmentation and peer matching', 'Application + Standard Census', True],

    # BUSINESS MODEL
    ['Business Model', 'Current e-commerce involvement', '', '', '', 'F2-Q3', 'Which best describes your current involvement?', '7 options: currently selling, sold, consulting, etc.', '', '', '',
     'Verify active seller status', 'Critical filter: separates active sellers from those who sold/exited. Members who no longer sell may need different programming or may not qualify', 'Application + Standard Census', False],
    ['Business Model', 'Business models used', 'F1-Q4', 'Which business models apply to you?', 'Private Label, Wholesale, Brand Mgmt, OEM', '', '', '', 'F3-Q8', 'Which business models apply?', 'Arb/Wholesale, PL, Brand Mgmt, OEM',
     'Segment members by business type', 'Different business models have different challenges. Enables targeted content, event tracks, and peer matching. PL sellers face different issues than wholesale', 'Application + Standard Census', False],
    ['Business Model', 'Strongest area of expertise', 'F1-Q3', 'What is your strongest area of expertise?', '', '', '', '', 'F3-Q6', 'What is your strongest area of expertise?', '',
     'Identify member value to community', 'Helps match members who need help with members who can help. Powers expert directories and speaker selection', 'Application only', False],
    ['Business Model', 'Validation of membership', '', '', '', '', '', '', 'F3-Q4', 'Are you a teacher/coach/vendor?', 'Group with sub-questions',
     'Screen for conflicts of interest', 'MDS is a seller community. Teachers/coaches/vendors may have commercial motives. Important for trust and keeping the group seller-focused', 'Application only', False],
    ['Business Model', 'Main niche/category', 'F1-Q17', 'What is your main niche?', '', '', '', '', 'F3-Q15', 'What is your main niche?', '',
     'Understand market focus', 'Enables niche-based matching (supplements sellers, electronics sellers, etc.). Helps plan relevant content and events', 'Application + Standard Census', False],

    # BRANDS & PRODUCTS
    ['Brands & Products', 'Number of brands', 'F1-Q6', 'How many brands do you currently have?', '0 to 5+', '', '', '', 'F3-Q7', 'How many brands do you have?', '0 to 5+',
     'Gauge business complexity', 'Multi-brand operators have different needs than single-brand. Indicates sophistication level and operational complexity', 'Application + Standard Census', False],
    ['Brands & Products', 'Brand name(s)', 'F1-Q7', 'What is the name of your brand(s)?', '', '', '', '', '', '', '',
     'Identify and verify brands', 'Enables staff to look up brands on Amazon/Shopify. Cross-reference with storefront data. Useful for member directory', 'Application + Standard Census', False],
    ['Brands & Products', 'Number of products (parent SKUs)', 'F1-Q8', 'How many products do you sell?', 'Max 99999', '', '', '', 'F3-Q11', 'How many parent listings?', '',
     'Measure catalog size', 'SKU count is a key business size indicator alongside revenue. 10-SKU vs 500-SKU businesses operate very differently', 'Application + Standard Census', False],
    ['Brands & Products', 'Product categories', 'F1-Q9', 'Which product categories?', '10 categories + Other', '', '', '', 'F3-Q14', 'Which product categories?', 'Same categories',
     'Segment by market vertical', 'Enables category-specific peer groups, content, and event tracks. Different categories have different margins, competition, and logistics', 'Application + Standard Census', False],
    ['Brands & Products', 'Products launched last year', '', '', '', 'F2-Q15', 'How many products launched last year?', '', 'F3-Q32', 'Products launched last year', '',
     'Measure growth velocity', 'Launch rate indicates business growth trajectory. High launchers vs steady-state operators need different advice', 'Standard Census + MDS-only', False],
    ['Brands & Products', 'New products planned this year', '', '', '', 'F2-Q16', 'How many new products planned?', '', 'F3-Q33', 'New products planned', '',
     'Forecast growth intent', 'Forward-looking metric. Helps identify members in growth mode who may need more support with launches, sourcing, etc.', 'Standard Census + MDS-only', False],
    ['Brands & Products', 'When started selling on Amazon', '', '', '', '', '', '', 'F3-Q13', 'Year started on Amazon', '2012-2021 dropdown',
     'Gauge experience level', 'One-time question. Veteran sellers (2012) vs newcomers (2020+) have very different perspectives and needs', 'Application only', False],

    # REVENUE
    ['Revenue', 'TTM total revenue', 'F1-Q15', 'TTM revenue (trailing 12 months)', 'Min 1M', '', '', '', 'F3-Q25', 'TTM revenue', '',
     'Verify membership qualification', 'THE key qualification metric. Must be $1M+ to be in MDS. Re-asked annually to confirm continued qualification', 'All 3 forms', False],
    ['Revenue', 'Projected FTM revenue', 'F1-Q14', 'Projected next 12 months revenue', 'Min 1M', '', '', '', 'F3-Q28', 'Projected revenue', '',
     'Forecast member growth trajectory', 'Shows if members are growing, flat, or declining. Helps MDS plan programming for the right stage of business', 'Application + Standard Census', False],
    ['Revenue', '% revenue off Amazon', '', '', '', '', '', '', 'F3-Q10', '% of revenue from non-Amazon', '',
     'Measure Amazon dependency', 'Members heavily dependent on Amazon have different risk profiles than diversified sellers. Informs content on diversification', 'Application + Standard Census', False],
    ['Revenue', 'Sales channels + revenue %', 'F1-Q5', 'Revenue % per channel', 'Matrix: Amazon, Shopify, Walmart, etc.', '', '', '', '', '', '',
     'Detailed channel mix breakdown', 'The SIMPLE revenue matrix. Shows WHERE revenue comes from. Critical for understanding member business structure', 'Application + Standard Census', False],
    ['Revenue', 'Revenue % by business category', '', '', '', 'F2-Q28', 'Revenue breakdown by biz category', 'Matrix with detailed categories', '', '', '',
     'Deep-dive revenue analysis', 'The COMPLEX revenue matrix. More granular than channel mix. Only valuable with large enough dataset to draw conclusions', 'MDS-only census', False],
    ['Revenue', 'Selling focus next 12 months', '', '', '', 'F2-Q17', 'Where focusing selling efforts?', 'Matrix: 10 channels with priority levels', '', '', '',
     'Predict channel shifts', 'Forward-looking channel strategy. Shows where sellers are heading (e.g. TikTok adoption, Walmart expansion)', 'MDS-only census', False],
    ['Revenue', 'Customer orders shipped (monthly)', '', '', '', 'F2-Q13', 'Monthly customer orders shipped', '', 'F3-Q27', 'Monthly orders shipped', '',
     'Measure operational volume', 'Revenue alone doesnt capture volume. High-volume low-ASP vs low-volume high-ASP are different operations', 'MDS-only census', False],

    # SUPPLY CHAIN
    ['Supply Chain', 'Avg cost per CBM', '', '', '', 'F2-Q10', 'Average cost per CBM', '', '', '', '',
     'Benchmark sourcing costs', 'Niche metric for importers. Useful if enough members respond, but historically low response rate', 'MDS-only census', False],
    ['Supply Chain', 'Avg production time', '', '', '', 'F2-Q11', 'Average production time (days)', '', '', '', '',
     'Benchmark manufacturing timelines', 'Helps members compare production lead times. Useful for supply chain planning discussions', 'MDS-only census', False],
    ['Supply Chain', 'Avg shipping time', '', '', '', 'F2-Q12', 'Average shipping time (days)', '', '', '', '',
     'Benchmark logistics timelines', 'Shipping times vary widely. Aggregate data helps members negotiate better rates or find alternatives', 'MDS-only census', False],
    ['Supply Chain', 'How do you source products?', '', '', '', 'F2-Q7', 'How do you source?', '', 'F3-Q29', 'How do you source products?', '',
     'Understand supply chain approach', 'Domestic vs overseas sourcing affects costs, timelines, and challenges. Enables sourcing-focused peer matching', 'Application + Standard Census', False],
    ['Supply Chain', 'Manufacturing locations', '', '', '', 'F2-Q8', 'Where do you manufacture?', 'Matrix with countries', 'F3-Q30', 'Manufacturing countries', '',
     'Map global supply chain', 'China vs India vs domestic manufacturing is a key strategic decision. Helps with tariff discussions and sourcing events', 'Standard Census + MDS-only', False],
    ['Supply Chain', 'Containers imported per year', '', '', '', 'F2-Q14', 'Containers imported per year', '', 'F3-Q31', 'Containers imported', '',
     'Measure import volume', 'Container count indicates scale of importing operations. Relevant for logistics cost benchmarking', 'MDS-only census', False],
    ['Supply Chain', 'Warehousing types used', '', '', '', 'F2-Q4', 'Warehousing types', 'FBA, 3PL, In-House', 'F3-Q16', 'Warehousing types', 'FBA, 3PL, SFP, In-House',
     'Understand fulfillment strategy', 'FBA-only vs 3PL vs in-house reveals operational maturity. Members using 3PL can share recommendations', 'Application + Standard Census', False],

    # ACQUISITIONS
    ['Acquisitions', 'Purchased a business before?', 'F1-Q10', 'Have you purchased an e-com business?', '', '', '', '', 'F3-Q44', 'Purchased a business?', '',
     'Identify acquisition-minded members', 'Growing segment interested in buying brands. Enables M&A focused events and matchmaking between buyers/sellers', 'Application + Standard Census', False],
    ['Acquisitions', 'Sold a business before?', 'F1-Q11', 'Have you sold an e-com business?', '', '', '', '', 'F3-Q45', 'Sold a business?', '',
     'Track exit activity', 'Members who have exited have different perspectives and needs. Also validates continued qualification', 'Application + Standard Census', False],
    ['Acquisitions', 'When did you sell?', 'F1-Q12', 'When did you sell?', '', '', '', '', '', '', '',
     'Timeline of exit', 'Context for the sale. Recent sellers may still qualify via new ventures. Old exits may mean member is no longer active', 'Standard Census', False],
    ['Acquisitions', 'Still have e-com revenue?', 'F1-Q13', 'Do you still have e-com revenue from new brands?', '', '', '', '', '', '', '',
     'Verify continued qualification', 'Critical: if they sold and have no new revenue, they may no longer qualify for MDS. Triggers re-verification', 'Standard Census', False],
    ['Acquisitions', 'Plan to sell in next 12 months?', 'F1-Q16', 'Do you plan to sell a brand?', '5 options', '', '', '', '', '', '',
     'Anticipate membership changes', 'Members planning to sell may exit the group. Also useful for M&A matchmaking with potential buyers in the group', 'Standard Census', False],

    # TEAM & OPS
    ['Team & Ops', 'W-2 employees count', '', '', '', 'F2-Q18', 'W-2 full-time employees', '', 'F3-Q36', 'Full-time W-2 employees', '',
     'Measure team size', 'Key operational metric. Businesses with 2 employees vs 50 operate very differently. Useful for peer matching by company size', 'Standard Census + MDS-only', False],
    ['Team & Ops', 'Part-time/1099 contractors', '', '', '', 'F2-Q19', 'Part-time or 1099 contractors', '', '', '', '',
     'Full picture of workforce', 'Many e-com businesses rely heavily on contractors. W-2 count alone understates true team size', 'MDS-only census', False],
    ['Team & Ops', 'VA/offshore employees', '', '', '', 'F2-Q20', 'VA or offshore employees', '', 'F3-Q37', 'VAs or offshore employees', '',
     'Understand labor strategy', 'VA usage is a key differentiator in e-com. Enables VA-focused content, tool recommendations, and hiring best practices sharing', 'Standard Census + MDS-only', False],
    ['Team & Ops', 'Staff locations', '', '', '', 'F2-Q21', 'Where is your team located?', '10 locations', 'F3-Q42', 'Staff locations', '6 locations',
     'Map global workforce', 'Helps understand if teams are domestic, offshore, or hybrid. Relevant for HR, management, and timezone discussions', 'Standard Census + MDS-only', False],
    ['Team & Ops', 'Team positions & roles', '', '', '', 'F2-Q22', 'Team positions & compensation', 'Matrix', 'F3-Q40', 'Team positions', 'MC multi',
     'Benchmark organizational structure', 'Shows which roles members have filled. Helps members planning their next hire see what peers have done', 'MDS-only census', False],
    ['Team & Ops', 'Other team positions', '', '', '', 'F2-Q23', 'Other positions not listed', '', 'F3-Q41', 'Other positions', '',
     'Capture unlisted roles', 'Catches roles not in the standard list. Reveals emerging positions (AI specialist, TikTok manager, etc.)', 'MDS-only census', False],
    ['Team & Ops', 'Team building advice', '', '', '', 'F2-Q24', 'Best team building advice', '', '', '', '',
     'Crowdsource member wisdom', 'Qualitative question for community value. Great answers can be shared as content. Low data analysis value', 'MDS-only census', False],
    ['Team & Ops', 'Use EOS/Traction?', '', '', '', 'F2-Q25', 'Do you use EOS/Traction?', 'Yes/No', 'F3-Q38', 'Use EOS?', '',
     'Gauge management framework adoption', 'EOS is popular among MDS members. Helps plan EOS-specific programming and peer groups', 'Standard Census + MDS-only', False],
    ['Team & Ops', 'EOS implementation level', '', '', '', 'F2-Q26', 'How far into EOS?', '', 'F3-Q39', 'EOS implementation', '',
     'Depth of EOS usage', 'Follow-up to above. Members early in EOS journey need different support than those fully implemented', 'Standard Census + MDS-only', False],

    # MARKETING & MANAGEMENT
    ['Marketing & Mgmt', 'Marketing initiatives handling', '', '', '', 'F2-Q5', 'How handle marketing?', 'Matrix: 8 channels (Amazon Ads, FB, Google, etc.)', '', '', '',
     'Map marketing operations', 'Shows what is in-house vs agency vs not done. Reveals gaps and opportunities for MDS content/tools recommendations', 'MDS-only census', False],
    ['Marketing & Mgmt', 'Business operations handling', '', '', '', 'F2-Q6', 'How handle business areas?', 'Matrix: 8 areas (product dev, CS, design, etc.)', '', '', '',
     'Map operational structure', 'Similar to marketing: in-house vs outsourced. Reveals operational maturity and where members might need help', 'MDS-only census', False],
    ['Marketing & Mgmt', 'Amazon PPC management', '', '', '', '', '', '', 'F3-Q17', 'Who manages Amazon PPC?', '',
     'Track PPC approach', 'Covered by F2-Q5 matrix more efficiently. Individual question approach is more granular but creates survey fatigue', 'Application (via matrix in census)', False],
    ['Marketing & Mgmt', 'External paid traffic', '', '', '', '', '', '', 'F3-Q18', 'Who manages external traffic?', '',
     'Track off-Amazon marketing', 'Same as above - covered by F2 matrix. Separate question only needed in application if matrix is too complex', 'Application (via matrix in census)', False],

    # TOOLS & SERVICES
    ['Tools & Services', 'Split testing tool', 'F1-Q28', 'What split testing tool do you use?', 'Multiple choices + Other', '', '', '', '', '', '',
     'Track tool adoption', 'Helps MDS negotiate group deals. Shows market share of tools among members. Can inform tool recommendation content', 'Standard Census', False],
    ['Tools & Services', 'PPC management service', 'F1-Q29', 'PPC management service/software', 'Multiple choices + Other', '', '', '', '', '', '',
     'Track PPC tool landscape', 'Same as above. PPC tools change fast. Annual tracking keeps recommendations current', 'Standard Census', False],
    ['Tools & Services', 'Reimbursement tool', 'F1-Q30', 'Reimbursement tool used', 'Multiple choices + Other', '', '', '', '', '', '',
     'Track reimbursement adoption', 'High-value category for partnerships. Most sellers leave money on the table with reimbursements', 'Standard Census', False],
    ['Tools & Services', '3PL management', 'F1-Q31', '3PL provider used', 'Choices + Other', '', '', '', '', '', '',
     'Map 3PL usage', 'Facilitates 3PL recommendations among members. Valuable for members transitioning from FBA-only', 'Standard Census', False],
    ['Tools & Services', 'HR/Recruitment services', 'F1-Q32', 'HR/Recruitment services used', 'Multiple choices + Other', '', '', '', '', '', '',
     'Track HR tool adoption', 'As teams grow, HR becomes critical. Helps members find vetted recruitment services used by peers', 'Standard Census', False],

    # GROWTH & STRATEGY
    ['Growth & Strategy', 'Goals for this year', 'F1-Q22', 'What are your goals this year?', '', '', '', '', 'F3-Q46', 'Goals for this year', '',
     'Understand member aspirations', 'Helps MDS align programming with what members want to achieve. Can track goal completion year over year', 'Application + Standard Census', False],
    ['Growth & Strategy', 'Biggest challenge this year', 'F1-Q23', 'What is your biggest challenge?', '', '', '', '', '', '', '',
     'Identify pain points', 'The #1 most actionable question for MDS programming. If 40% say hiring, run hiring-focused events', 'Standard Census', False],
    ['Growth & Strategy', 'How plan to grow business', 'F1-Q25', 'How do you plan to grow in next 12 months?', '', '', '', '', '', '', '',
     'Map growth strategies', 'Qualitative insights on member strategies. Good for content creation and identifying common approaches', 'Standard Census', False],
    ['Growth & Strategy', 'Best thing that worked', 'F1-Q26', 'Best thing that worked in last 12 months?', '', '', '', '', '', '', '',
     'Crowdsource wins', 'High-value for community. Best answers become shareable content. Shows what is actually working for sellers', 'Standard Census', False],
    ['Growth & Strategy', 'Most impactful service/software', 'F1-Q27', 'Most impactful service or software?', '', '', '', '', '', '', '',
     'Track high-value tools', 'Peer-validated recommendations are gold. This question yields the most actionable tool suggestions', 'Standard Census', False],
    ['Growth & Strategy', 'Competitive advantage', 'F1-Q24', 'What is your competitive advantage?', 'MC multi + Other', '', '', '', '', '', '',
     'Understand member moats', 'Reveals how members differentiate. Helps segment by strategy (price vs brand vs IP vs operations)', 'Standard Census', False],
    ['Growth & Strategy', 'Industries spending 20%+ time', 'F1-Q21', 'What industries do you spend time in?', '10 options + Other', '', '', '', '', '', '',
     'Track diversification beyond e-com', 'Many members invest in real estate, crypto, SaaS, etc. Enables cross-industry networking and content', 'Standard Census', False],
    ['Growth & Strategy', 'Business change YoY', '', '', '', 'F2-Q27', 'How business changed year over year', 'Matrix', '', '', '',
     'Track business trajectory', 'Quantitative YoY comparison across metrics. Valuable if response rate is high enough for statistical significance', 'MDS-only census', False],
    ['Growth & Strategy', 'New business ventures', '', '', '', 'F2-Q29', 'Any new business ventures?', '', '', '', '',
     'Track member diversification', 'Members starting new ventures may need different support. Also identifies potential mentors for new verticals', 'MDS-only census', False],
    ['Growth & Strategy', 'Current investments', '', '', '', 'F2-Q30', 'Current investments outside e-com', '', '', '', '',
     'Understand wealth allocation', 'Sensitive but valuable. Many members invest in real estate, startups. Enables investor-focused networking', 'MDS-only census', False],

    # COMMUNITY
    ['Community', 'Other groups/communities', 'F1-Q20', 'Other knowledge bases or groups?', '', '', '', '', 'F3-Q43', 'Other groups you belong to', '',
     'Map competitive landscape', 'Shows which other communities members value. Helps MDS differentiate and understand where members get value elsewhere', 'Application + Standard Census', False],
    ['Community', 'What have you been up to?', 'F1-Q48', 'What have you been up to?', '', '', '', '', '', '', '',
     'Personal check-in', 'Open-ended relationship builder. Shows MDS cares about the person, not just the business. Great for community feel', 'Standard Census', False],

    # MDS FEEDBACK
    ['MDS Feedback', 'Rank member benefits', 'F1-Q33', 'Rank these member benefits', 'Ranking: 11 items', '', '', '', '', '', '',
     'Prioritize MDS offerings', 'Shows which benefits members value most. Directly informs budget and resource allocation for programs', 'Standard Census', False],
    ['MDS Feedback', 'Virtual call topics wanted', 'F1-Q34', 'What virtual call topics?', '', '', '', '', '', '', '',
     'Plan event content', 'Direct input for event programming. Low effort question, high value output', 'Standard Census', False],
    ['MDS Feedback', 'Visual branding resonance', 'F1-Q35', 'Does MDS branding resonate?', 'Opinion scale', '', '', '', '', '', '',
     'Evaluate brand perception', 'Track if rebrand efforts land. Low priority unless MDS is actively rebranding', 'Standard Census', False],
    ['MDS Feedback', 'Branding improvement ideas', 'F1-Q36', 'How to improve branding?', '', '', '', '', '', '', '',
     'Crowdsource brand feedback', 'Follow-up to above. Only useful if branding score is low', 'Standard Census', False],
    ['MDS Feedback', 'MDS community impact', 'F1-Q37', 'How has MDS impacted you?', '', '', '', '', '', '', '',
     'Collect testimonials & measure value', 'Best responses become marketing testimonials. Also measures perceived ROI of membership', 'Standard Census', False],
    ['MDS Feedback', 'NPS score', 'F1-Q38', 'How likely to recommend MDS?', 'Opinion scale 0-10', '', '', '', '', '', '',
     'Track member satisfaction', 'Industry standard loyalty metric. Track YoY to measure community health. Scores below 7 trigger concern', 'Standard Census', False],
    ['MDS Feedback', 'How to make MDS better', 'F1-Q39', 'How can MDS be better for you?', '', '', '', '', 'F3-Q31', 'How can MDS improve?', '',
     'Direct improvement feedback', 'Most actionable feedback question. Answers drive the MDS roadmap', 'Standard Census', False],
    ['MDS Feedback', 'Know friends who qualify?', 'F1-Q40', 'Know friends whod qualify for MDS?', 'Yes/No', '', '', '', 'F3-Q47', 'Know qualified friends?', '',
     'Drive referrals', 'Membership growth channel. Yes answers should trigger a referral follow-up workflow', 'Application + Standard Census', False],
    ['MDS Feedback', 'UX rating of MDS systems', 'F1-Q49', 'Rate UX of MDS systems', 'Opinion scale', '', '', '', '', '', '',
     'Track tech satisfaction', 'Measures satisfaction with MDS platform, tools, app. Low scores flag tech investment needs', 'Standard Census', False],
    ['MDS Feedback', 'Tech enhancement areas', 'F1-Q50', 'What tech areas need improvement?', '', '', '', '', '', '', '',
     'Prioritize tech roadmap', 'Follow-up to UX rating. Specific suggestions for tech team', 'Standard Census', False],

    # PROGRAMS
    ['MDS Programs', 'MDS Chapters involvement', 'F1-Q43', 'Are you involved in a Chapter?', 'Group: 3 sub-Qs', '', '', '', '', '', '',
     'Measure Chapter engagement', 'Track Chapter participation rate, satisfaction, and improvement ideas', 'Standard Census', False],
    ['MDS Programs', 'MDS Programs participation', 'F1-Q44', 'Have you participated in Programs?', 'Group: 3 sub-Qs', '', '', '', '', '', '',
     'Measure Program engagement', 'Same as Chapters but for Programs (Masterminds, Courses, etc.)', 'Standard Census', False],
    ['MDS Programs', 'MDS Squads involvement', 'F1-Q45', 'Are you in a Squad?', 'Yes/No + rating + feedback', '', '', '', '', '', '',
     'Measure Squad engagement', 'Squads are a key differentiator. Track participation and satisfaction to improve the program', 'Standard Census', False],

    # PERSONAL
    ['Personal', 'Number of kids', 'F1-Q41', 'How many kids do you have?', '', '', '', '', '', '', '',
     'Personal connection building', 'Enables family-oriented events, kid-friendly meetups. Shows MDS cares about whole person', 'Standard Census', False],
    ['Personal', 'Kids age ranges', 'F1-Q42', 'What are your kids age ranges?', '', '', '', '', '', '', '',
     'Family stage segmentation', 'Parents of toddlers vs teenagers have different availability and interests for events', 'Standard Census', False],

    # LEGAL
    ['Legal', 'NDA/Privacy agreement', '', '', '', '', '', '', 'F3-Q49', 'NDA/Privacy/Terms confirmation', 'Group: confirm + date',
     'Legal compliance', 'Required for membership. Protects confidential information shared within the group', 'Application only', False],

    # NEW: PREQUALIFICATION
    ['Channels', 'Closed channel prequalification', '', '', '', '', '', '', '', '', '',
     'Auto-qualify for private channels', 'Remove friction: instead of separate application for closed channels (e.g. Supplements, 8-Figure), qualify members based on census answers automatically', 'Application + Standard Census', True],
]

# ====== SHEET 1: STRATEGIC REVIEW ======
ws = wb.active
ws.title = 'Strategic Review'

headers = ['Category', 'Topic / Question',
    'Form 1 (Standard Census) Ref', 'Form 1 Question Text', 'Form 1 Subtext',
    'Form 2 (MDS-only) Ref', 'Form 2 Question Text', 'Form 2 Subtext',
    'Form 3 (Application) Ref', 'Form 3 Question Text', 'Form 3 Subtext',
    'PURPOSE', 'WHY ARE WE ASKING THIS?', 'WHICH FORM(S) SHOULD HAVE THIS?']

col_widths = [18, 32, 10, 40, 30, 10, 40, 30, 10, 40, 30, 35, 55, 30]

# Legend
ws.append(['LEGEND:', '', 'Green highlighted rows = NEW QUESTIONS to add'])
ws['C1'].fill = NEW_FILL
ws['C1'].font = Font(bold=True, size=11)
ws['A1'].font = Font(bold=True, size=12)
ws.append([])

# Headers
ws.append(headers)
for col_idx, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width
    cell = ws.cell(row=3, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='middle')
    cell.border = BORDER

# Data rows
last_cat = ''
for i, r in enumerate(master_rows):
    is_new = r[14]
    row_data = r[:14]
    if is_new:
        row_data[1] = 'NEW: ' + row_data[1]
    ws.append(row_data)
    row_num = ws.max_row

    for col_idx in range(1, 15):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.alignment = WRAP
        cell.border = BORDER

        if is_new:
            cell.fill = NEW_FILL
            cell.font = NEW_FONT
        elif r[0] != last_cat and col_idx == 1:
            cell.fill = CAT_FILL
            cell.font = Font(bold=True)
        elif i % 2 == 0 and not is_new:
            cell.fill = ALT_FILL

    # Color the Purpose/Why/Form columns
    if not is_new:
        ws.cell(row=row_num, column=12).font = PURPOSE_FONT
        ws.cell(row=row_num, column=13).font = WHY_FONT
        ws.cell(row=row_num, column=14).font = FORM_FONT

    last_cat = r[0]

ws.freeze_panes = 'C4'
ws.auto_filter.ref = f'A3:N{ws.max_row}'

# ====== SHEET 2: SUMMARY ======
ws2 = wb.create_sheet('Summary & Recommendations')
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 60
ws2.column_dimensions['C'].width = 40

ws2.append(['MDS CENSUS 2026 - STRATEGIC REVIEW SUMMARY'])
ws2['A1'].font = Font(bold=True, size=16, color='1B2A4A')
ws2.merge_cells('A1:C1')
ws2.append([])

ws2.append(['THE 3 FORMS & THEIR PURPOSES'])
ws2['A3'].font = Font(bold=True, size=14, color='1565C0')
ws2.append([])

for val in ['Form', 'Purpose', 'When Used']:
    pass
row = ws2.append(['Form', 'Purpose', 'When Used'])
for col in range(1, 4):
    c = ws2.cell(row=5, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = BORDER

form_info = [
    ['Form 3: Application', 'Verify new members & collect baseline info. One-time questions (education, when started) plus business data for qualification.', 'New member applies to join MDS'],
    ['Form 1: Standard Census', 'Update changing data annually. Skips one-time questions. Revenue, tools, strategy, MDS feedback, programs.', 'Membership renewal/expiration'],
    ['Form 2: MDS-only Census', 'Deep-dive operations, financials, team. Nuanced questions requiring lookup. Historically low response (~50).', 'Sent to existing members annually (optional)'],
]
for fi in form_info:
    ws2.append(fi)
    for col in range(1, 4):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.alignment = WRAP
        c.border = BORDER

ws2.append([])
ws2.append([])
ws2.append(['NEW QUESTIONS TO ADD (from ticket)'])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14, color='2E7D32')
ws2.append([])

for val in ['New Question', 'Purpose', 'Add to Which Form(s)']:
    pass
ws2.append(['New Question', 'Purpose', 'Add to Which Form(s)'])
for col in range(1, 4):
    c = ws2.cell(row=ws2.max_row, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = BORDER

new_qs = [
    ['What is your business name?', 'Official business ID for verification and records', 'Application + Standard Census'],
    ['What is your website?', 'Verify business legitimacy and size', 'Application + Standard Census'],
    ['What is your Amazon storefront name/URL?', 'Primary verification of Amazon presence and revenue', 'Application + Standard Census'],
    ['What is your TikTok shop name/code?', 'Track TikTok commerce adoption', 'Application + Standard Census'],
    ['Revenue screenshot upload', 'Verify self-reported revenue with hard evidence', 'Application + Standard Census'],
    ['TikTok annual revenue', 'Track fastest-growing channel separately', 'Application + Standard Census'],
    ['Shopify annual revenue', 'Track DTC revenue - indicator of brand strength', 'Application + Standard Census'],
    ['Closed channel prequalification', 'Auto-qualify for private channels based on census answers', 'Application + Standard Census'],
]
for nq in new_qs:
    ws2.append(nq)
    for col in range(1, 4):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.fill = NEW_FILL
        c.font = NEW_FONT
        c.alignment = WRAP
        c.border = BORDER

ws2.append([])
ws2.append([])
ws2.append(['STATISTICS'])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=14, color='E65100')
ws2.append([])

ws2.append(['Metric', 'Value', ''])
for col in range(1, 4):
    c = ws2.cell(row=ws2.max_row, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.border = BORDER

new_count = sum(1 for r in master_rows if r[14])
existing_count = sum(1 for r in master_rows if not r[14])
app_only = sum(1 for r in master_rows if r[13] == 'Application only')
std_census = sum(1 for r in master_rows if 'Standard Census' in r[13])
mds_only = sum(1 for r in master_rows if 'MDS-only' in r[13])

stats = [
    ['Total existing question topics', str(existing_count)],
    ['New questions to add', str(new_count)],
    ['Form 1 (Standard Census) questions', '50'],
    ['Form 2 (MDS-only Census) questions', '32'],
    ['Form 3 (Application) questions', '49'],
    ['Recommended for Application only', str(app_only)],
    ['Recommended for Standard Census (incl combos)', str(std_census)],
    ['Recommended for MDS-only (incl combos)', str(mds_only)],
]
for s in stats:
    ws2.append([s[0], s[1], ''])
    for col in range(1, 4):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.alignment = WRAP
        c.border = BORDER
    ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True, size=12)

# Save
output_path = '/Users/Born/Scorecard/MDS_Census_2026_Strategic_Review.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
